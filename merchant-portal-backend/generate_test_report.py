"""
Parses Maven Surefire XML reports and generates an Excel test report.
Usage: python generate_test_report.py
"""
import xml.etree.ElementTree as ET
import glob
import os

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    import subprocess, sys
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

REPORT_DIR = os.path.join("target", "surefire-reports")
OUTPUT_FILE = "test-report.xlsx"

# Styles
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
PASS_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
FAIL_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
SKIP_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
PASS_FONT = Font(name="Calibri", color="006100")
FAIL_FONT = Font(name="Calibri", color="9C0006")
SKIP_FONT = Font(name="Calibri", color="9C6500")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)

wb = Workbook()

# ── Sheet 1: Summary ──
ws_summary = wb.active
ws_summary.title = "Summary"

# ── Sheet 2: All Test Cases ──
ws_detail = wb.create_sheet("Test Cases")

detail_headers = ["#", "Test Class", "Test Method", "Status", "Time (s)", "Failure Message"]
for col, h in enumerate(detail_headers, 1):
    cell = ws_detail.cell(row=1, column=col, value=h)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal="center")
    cell.border = THIN_BORDER

# Parse all XML reports
xml_files = sorted(glob.glob(os.path.join(REPORT_DIR, "TEST-*.xml")))

total_tests = 0
total_pass = 0
total_fail = 0
total_error = 0
total_skip = 0
total_time = 0.0
suite_rows = []  # for summary sheet
detail_row = 2

for xml_file in xml_files:
    tree = ET.parse(xml_file)
    root = tree.getroot()

    suite_name = root.attrib.get("name", "Unknown")
    tests = int(root.attrib.get("tests", 0))
    failures = int(root.attrib.get("failures", 0))
    errors = int(root.attrib.get("errors", 0))
    skipped = int(root.attrib.get("skipped", 0))
    time_taken = float(root.attrib.get("time", 0))
    passed = tests - failures - errors - skipped

    total_tests += tests
    total_pass += passed
    total_fail += failures
    total_error += errors
    total_skip += skipped
    total_time += time_taken

    suite_status = "PASS" if (failures == 0 and errors == 0) else "FAIL"
    suite_rows.append((suite_name, tests, passed, failures, errors, skipped, f"{time_taken:.3f}", suite_status))

    # Detail rows
    for tc in root.findall("testcase"):
        tc_class = tc.attrib.get("classname", "")
        tc_name = tc.attrib.get("name", "")
        tc_time = float(tc.attrib.get("time", 0))
        failure_msg = ""

        failure = tc.find("failure")
        error = tc.find("error")
        skip = tc.find("skipped")

        if failure is not None:
            status = "FAIL"
            failure_msg = failure.attrib.get("message", "")[:200]
        elif error is not None:
            status = "ERROR"
            failure_msg = error.attrib.get("message", "")[:200]
        elif skip is not None:
            status = "SKIPPED"
        else:
            status = "PASS"

        row_data = [detail_row - 1, tc_class, tc_name, status, f"{tc_time:.3f}", failure_msg]
        for col, val in enumerate(row_data, 1):
            cell = ws_detail.cell(row=detail_row, column=col, value=val)
            cell.border = THIN_BORDER
            if col == 4:  # Status column
                if status == "PASS":
                    cell.fill = PASS_FILL
                    cell.font = PASS_FONT
                elif status in ("FAIL", "ERROR"):
                    cell.fill = FAIL_FILL
                    cell.font = FAIL_FONT
                elif status == "SKIPPED":
                    cell.fill = SKIP_FILL
                    cell.font = SKIP_FONT
                cell.alignment = Alignment(horizontal="center")
        detail_row += 1

# ── Build Summary Sheet ──
summary_headers = ["Test Suite", "Total", "Passed", "Failed", "Errors", "Skipped", "Time (s)", "Status"]
for col, h in enumerate(summary_headers, 1):
    cell = ws_summary.cell(row=1, column=col, value=h)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal="center")
    cell.border = THIN_BORDER

for i, row_data in enumerate(suite_rows, 2):
    for col, val in enumerate(row_data, 1):
        cell = ws_summary.cell(row=i, column=col, value=val)
        cell.border = THIN_BORDER
        if col == 8:  # Status
            if val == "PASS":
                cell.fill = PASS_FILL
                cell.font = PASS_FONT
            else:
                cell.fill = FAIL_FILL
                cell.font = FAIL_FONT
            cell.alignment = Alignment(horizontal="center")

# Totals row
totals_row = len(suite_rows) + 2
totals = ["TOTAL", total_tests, total_pass, total_fail, total_error, total_skip, f"{total_time:.3f}",
          "PASS" if (total_fail == 0 and total_error == 0) else "FAIL"]
for col, val in enumerate(totals, 1):
    cell = ws_summary.cell(row=totals_row, column=col, value=val)
    cell.font = Font(name="Calibri", bold=True, size=11)
    cell.fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
    cell.border = THIN_BORDER
    if col == 8:
        if val == "PASS":
            cell.fill = PASS_FILL
            cell.font = Font(name="Calibri", bold=True, color="006100")
        else:
            cell.fill = FAIL_FILL
            cell.font = Font(name="Calibri", bold=True, color="9C0006")
        cell.alignment = Alignment(horizontal="center")

# Auto-width columns
for ws in [ws_summary, ws_detail]:
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 3, 60)

wb.save(OUTPUT_FILE)
print(f"\n✅ Test report generated: {os.path.abspath(OUTPUT_FILE)}")
print(f"   Total: {total_tests} | Passed: {total_pass} | Failed: {total_fail} | Errors: {total_error} | Skipped: {total_skip}")

